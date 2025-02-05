import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Folder containing JSON files
script_dir = os.path.dirname(__file__)  
json_folder = os.path.join(script_dir, "JSON_Output_vision_validation")
output_comparison_csv = "validation_invoice_comparison.csv"
output_merged_xlsx = "validation_merged_invoices.xlsx"

# Initialize storage structures
invoice_data = {}
merged_data = {}

# Process each JSON file in the folder
for filename in os.listdir(json_folder):
    if filename.endswith(".json"):
        file_path = os.path.join(json_folder, filename)

        # Extract the invoice base name from the filename (before "_attempt_")
        invoice_base = filename.split("_attempt_")[0].lower()  # Convert to lowercase

        # Read JSON file
        with open(file_path, "r", encoding="utf-8") as file:
            try:
                json_data = json.load(file)

                # Ensure it's a list of entries
                if not isinstance(json_data, list):
                    json_data = [json_data]

                attempt_num = filename.split("_attempt_")[-1].split(".json")[0]  # Extract attempt number

                for entry in json_data:
                    energy_type = entry.get("energy_type", "Unknown").lower()  # Convert to lowercase

                    # Create unique invoice + energy type key (using extracted filename)
                    invoice_key = f"{invoice_base}__{energy_type}"

                    # Ensure invoice_key exists in both storage structures
                    if invoice_key not in invoice_data:
                        invoice_data[invoice_key] = {"invoice": invoice_base, "energy_type": energy_type}
                    if invoice_key not in merged_data:
                        merged_data[invoice_key] = {"invoice": invoice_base, "energy_type": energy_type}

                    # Store attempt data (convert values to lowercase where applicable)
                    for field in ["location", "usage_start_date", "usage_end_date", "energy_volume", "energy_units", "cost_amount", "currency"]:
                        value = entry.get(field, "N/A")

                        # Convert strings to lowercase, keep numbers unchanged
                        if isinstance(value, str):
                            value = value.lower()

                        attempt_field = f"{field}_{attempt_num}"
                        invoice_data[invoice_key][attempt_field] = value

                        # Store values for merging
                        if field not in merged_data[invoice_key]:
                            merged_data[invoice_key][field] = []
                        merged_data[invoice_key][field].append(value)

            except json.JSONDecodeError:
                print(f"Error decoding JSON in file: {filename}")

# Convert invoice_data to DataFrame (Detailed comparison output)
df_comparison = pd.DataFrame.from_dict(invoice_data, orient="index")

# Function to merge and flag inconsistencies
def merge_and_flag(row):
    flagged = []
    merged_row = {"invoice": row["invoice"], "energy_type": row["energy_type"]}

    for field in ["location", "usage_start_date", "usage_end_date", "energy_volume", "energy_units", "cost_amount", "currency"]:
        if field in row:
            values = set(row[field])  # Get unique values, ignoring capitalization
            if len(values) > 1:  # If inconsistent, flag it
                flagged.append(field)
                merged_row[field] = ", ".join(map(str, values))  # Store all variations in one cell
            else:
                merged_row[field] = next(iter(values))  # Store single value if consistent

    merged_row["Inconsistencies"] = ", ".join(flagged) if flagged else "Consistent"
    return merged_row

# Apply merge and flagging function
df_merged = pd.DataFrame.from_dict(merged_data, orient="index").apply(merge_and_flag, axis=1, result_type="expand")


# Save merged output as an Excel file with colors
wb = Workbook()
ws = wb.active
ws.title = "Merged Invoices"

# Define styles
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Light red for inconsistent fields
default_fill = PatternFill(fill_type=None)  # No fill for consistent fields

# Write headers
headers = ["invoice", "energy_type", "location", "usage_start_date", "usage_end_date", "energy_volume", "energy_units", "cost_amount", "currency", "Inconsistencies"]
ws.append(headers)

# Write data with conditional formatting
for row_idx, row in df_merged.iterrows():
    row_values = [row["invoice"], row["energy_type"], row["location"], row["usage_start_date"],
                  row["usage_end_date"], row["energy_volume"], row["energy_units"],
                  row["cost_amount"], row["currency"], row["Inconsistencies"]]

    ws.append(row_values)
    for col_idx, field in enumerate(headers[2:], start=3):  # Skip invoice & energy_type columns
        cell = ws.cell(row=ws.max_row, column=col_idx)
        if field in row["Inconsistencies"]:  # Apply red color if inconsistent
            cell.fill = red_fill
        else:
            cell.fill = default_fill  # No fill if consistent

# Save the Excel file
wb.save(output_merged_xlsx)

print(f"✅ Comparison CSV saved as: {output_comparison_csv}")
print(f"✅ Merged Excel file with color coding saved as: {output_merged_xlsx}")
